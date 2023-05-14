import re

from openpyxl import load_workbook
from unidecode import unidecode


def filter_excel_file():
    book = load_workbook("./CNAE_Details.xlsx")

    working_sheet = book["CNAE"]

    index_name_columns = [1, 3, 5, 7, 9]
    index_code_columns = [2, 4, 6, 8]

    for index, row in enumerate(working_sheet.iter_rows(min_row=2, max_row=working_sheet.max_row), start=2):
        for name_index in index_name_columns:
            column_value = row[name_index].value

            if not column_value:
                break

            column_value = column_value.lower()

            row[name_index].value = unidecode(column_value)

        for code_index in index_code_columns:
            column_value = str(row[code_index].value)

            if not column_value:
                break

            column_value = re.findall('[0-9]+', column_value)

            row[code_index].value = "".join(column_value)

    book.save("./CNAE_Details.xlsx")


filter_excel_file()